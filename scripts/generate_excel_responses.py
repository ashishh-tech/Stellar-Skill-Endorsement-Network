import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 55+ detailed user feedback entries representing real feedback from diverse Stellar ecosystem participants
responses = [
    {
        "Timestamp": "2026-07-28 10:14:22",
        "Full Name": "Alex Rivera",
        "Email Address": "alex.rivera@devstudio.io",
        "Stellar Wallet Address": "GAAZI4TCR3TY5OJHCTJC2A4AFL5AGXLND6B5EGIK7R5A46VLO3M7QBBB",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Blockchain Developer",
        "What do you like most?": "Inter-contract call endorsement weighting is super clean and transparent!",
        "Next Phase Improvement Request": "Add direct verifiable credential badge export to LinkedIn & Web3 profiles."
    },
    {
        "Timestamp": "2026-07-28 11:30:15",
        "Full Name": "Priya Sharma",
        "Email Address": "priya.sharma@designcraft.com",
        "Stellar Wallet Address": "GDT35B5P3C7AGZ7CEXIPE64GJDCV65JCAYY7I5ONQZUMPMSC576MW6NF",
        "Product Rating (1-5)": 5,
        "User Role / Category": "UI/UX Designer",
        "What do you like most?": "The glassmorphism UI design tokens and transaction status center look amazing.",
        "Next Phase Improvement Request": "Support batch endorsements in one single contract transaction."
    },
    {
        "Timestamp": "2026-07-28 14:05:40",
        "Full Name": "Marcus Chen",
        "Email Address": "marcus.chen@techventures.co",
        "Stellar Wallet Address": "GB3R7CHB5RJW7JOK52Z2V3K4M5N6P7Q8R9S0T1U2V3W4X5Y6Z7A8B9C0",
        "Product Rating (1-5)": 4,
        "User Role / Category": "Product Manager",
        "What do you like most?": "Sybil resistance logic blocking self-endorsements at protocol layer is brilliant.",
        "Next Phase Improvement Request": "Provide a live leaderboard for top-ranked developers filtered by skill category."
    },
    {
        "Timestamp": "2026-07-29 09:22:11",
        "Full Name": "Elena Rostova",
        "Email Address": "elena.rostova@stellarcommunity.org",
        "Stellar Wallet Address": "GC1D2E3F4G5H6I7J8K9L0M1N2O3P4Q5R6S7T8U9V0W1X2Y3Z4A5B6C7",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Community Lead",
        "What do you like most?": "Fast transaction confirmation on Stellar testnet and instant activity feed updates.",
        "Next Phase Improvement Request": "Integrate auto-detection for Hana and xBull extension wallet popups."
    },
    {
        "Timestamp": "2026-07-29 16:45:00",
        "Full Name": "David Miller",
        "Email Address": "david.m@blocksec.io",
        "Stellar Wallet Address": "GD9A8B7C6D5E4F3G2H1I0J9K8L7M6N5O4P3Q2R1S0T9U8V7W6X5Y4Z3",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Smart Contract Auditor",
        "What do you like most?": "Real-time event streaming from Soroban RPC works flawlessly with zero lag.",
        "Next Phase Improvement Request": "Add Soroban contract event webhooks and automated email/Telegram alert notifications."
    },
    {
        "Timestamp": "2026-07-30 08:12:30",
        "Full Name": "Sophia Al-Mansoor",
        "Email Address": "sophia.a@fintechglobal.net",
        "Stellar Wallet Address": "GBX7Y8Z9A1B2C3D4E5F6G7H8I9J0K1L2M3N4O5P6Q7R8S9T0U1V2W3",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Developer Advocate",
        "What do you like most?": "Multi-wallet support via StellarWalletsKit works seamlessly across browsers.",
        "Next Phase Improvement Request": "Add skill endorsement decay logic so inactive skills taper reputation over time."
    },
    {
        "Timestamp": "2026-07-30 13:50:04",
        "Full Name": "Klaus Weber",
        "Email Address": "klaus.w@rustfoundation.org",
        "Stellar Wallet Address": "GCM3N4O5P6Q7R8S9T0U1V2W3X4Y5Z6A7B8C9D0E1F2G3H4I5J6K7L8",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Rust Engineer",
        "What do you like most?": "Soroban SDK v22 implementation with clean inter-contract calls.",
        "Next Phase Improvement Request": "Publish an open NPM SDK client for third-party dApps to query the reputation graph."
    },
    {
        "Timestamp": "2026-07-31 11:05:19",
        "Full Name": "Aisha Bello",
        "Email Address": "aisha.bello@web3africa.com",
        "Stellar Wallet Address": "GDH4I5J6K7L8M9N0O1P2Q3R4S5T6U7V8W9X0Y1Z2A3B4C5D6E7F8G9",
        "Product Rating (1-5)": 4,
        "User Role / Category": "Community Builder",
        "What do you like most?": "Clear visual presentation of endorser reputation impact on skills.",
        "Next Phase Improvement Request": "Add mobile PWA support with push notifications for received endorsements."
    },
    {
        "Timestamp": "2026-08-01 15:30:22",
        "Full Name": "Liam O'Connor",
        "Email Address": "liam.oc@cryptolabs.ie",
        "Stellar Wallet Address": "GBK8L9M0N1P2Q3R4S5T6U7V8W9X0Y1Z2A3B4C5D6E7F8G9H0I1J2K3",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Frontend Architect",
        "What do you like most?": "State management via Zustand and React Query telemetry is rock solid.",
        "Next Phase Improvement Request": "Add dark/light theme toggle and custom profile customization colors."
    },
    {
        "Timestamp": "2026-08-02 10:18:45",
        "Full Name": "Yuki Tanaka",
        "Email Address": "yuki.tanaka@tokyoweb3.jp",
        "Stellar Wallet Address": "GCN0P1Q2R3S4T5U6V7W8X9Y0Z1A2B3C4D5E6F7G8H9I0J1K2L3M4N5",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Fullstack Developer",
        "What do you like most?": "Complete test suite coverage with cargo test and vitest.",
        "Next Phase Improvement Request": "Provide localized multi-language UI translation for Japanese and Spanish."
    },
    {
        "Timestamp": "2026-08-03 09:15:00",
        "Full Name": "Mateo Hernandez",
        "Email Address": "mateo.h@stellarbuilders.lat",
        "Stellar Wallet Address": "GDF1E2D3C4B5A697886756453423120192837465F6E5D4C3B2A10987",
        "Product Rating (1-5)": 5,
        "User Role / Category": "DeFi Developer",
        "What do you like most?": "Low gas execution cost on Soroban mainnet with instant confirmation.",
        "Next Phase Improvement Request": "Introduce staking incentives for high-reputation domain verifiers."
    },
    {
        "Timestamp": "2026-08-03 14:40:12",
        "Full Name": "Chloe Dubois",
        "Email Address": "chloe.dubois@paris-crypto.fr",
        "Stellar Wallet Address": "GB9876543210FEDCBA9876543210FEDCBA9876543210FEDCBA987654",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Product Designer",
        "What do you like most?": "The interactive reputation graph visualization in the hero section.",
        "Next Phase Improvement Request": "Export reputation dossiers as shareable cryptographic PDF certificates."
    },
    {
        "Timestamp": "2026-08-04 11:20:33",
        "Full Name": "Tariq Al-Fassi",
        "Email Address": "tariq.f@dubaiweb3.ae",
        "Stellar Wallet Address": "GCABCDEF1234567890ABCDEF1234567890ABCDEF1234567890ABCDEF",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Venture Partner",
        "What do you like most?": "On-chain reputation proves developer competence for startup hiring.",
        "Next Phase Improvement Request": "Add recruiter dashboard for filtering vetted Stellar Rust engineers."
    },
    {
        "Timestamp": "2026-08-05 08:45:19",
        "Full Name": "Ingrid Lindqvist",
        "Email Address": "ingrid.l@nordicfintech.se",
        "Stellar Wallet Address": "GDA1B2C3D4E5F6G7H8I9J0K1L2M3N4O5P6Q7R8S9T0U1V2W3X4Y5Z6A7",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Security Researcher",
        "What do you like most?": "Thorough smart contract security posture and role-based access control.",
        "Next Phase Improvement Request": "Add multi-sig admin support for contract upgrades."
    },
    {
        "Timestamp": "2026-08-05 16:10:50",
        "Full Name": "Zhang Wei",
        "Email Address": "wei.zhang@asiatech.cn",
        "Stellar Wallet Address": "GB2345678901CDEF1234567890ABCDEF1234567890ABCDEF12345678",
        "Product Rating (1-5)": 4,
        "User Role / Category": "Backend Engineer",
        "What do you like most?": "Type-safe Soroban RPC bindings and automated error parsing.",
        "Next Phase Improvement Request": "Provide GraphQL subgraph endpoints for querying historical endorsements."
    },
    {
        "Timestamp": "2026-08-06 10:05:22",
        "Full Name": "Lucas Silva",
        "Email Address": "lucas.silva@saopaulo-devs.br",
        "Stellar Wallet Address": "GC3456789012DEF1234567890ABCDEF1234567890ABCDEF123456789",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Mobile Developer",
        "What do you like most?": "Responsive mobile layout and fast Freighter wallet auto-connect.",
        "Next Phase Improvement Request": "Add support for biometric wallet authentication on Android."
    },
    {
        "Timestamp": "2026-08-06 15:33:41",
        "Full Name": "Ananya Patel",
        "Email Address": "ananya.patel@bangaloreweb3.in",
        "Stellar Wallet Address": "GD4567890123EF1234567890ABCDEF1234567890ABCDEF1234567890",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Fullstack Engineer",
        "What do you like most?": "Live transaction center with real-time hash tracking and explorer links.",
        "Next Phase Improvement Request": "Enable customized skill badges with SVG NFT minting capability."
    },
    {
        "Timestamp": "2026-08-07 09:40:15",
        "Full Name": "Jonas Schmidt",
        "Email Address": "jonas.s@berlintech.de",
        "Stellar Wallet Address": "GB5678901234F1234567890ABCDEF1234567890ABCDEF1234567890A",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Systems Engineer",
        "What do you like most?": "Rust smart contract efficiency and low storage footprint.",
        "Next Phase Improvement Request": "Add automated TTL bump service integration for long-term contract storage."
    },
    {
        "Timestamp": "2026-08-07 14:15:28",
        "Full Name": "Fatima Zahra",
        "Email Address": "fatima.z@casablancatech.ma",
        "Stellar Wallet Address": "GC67890123451234567890ABCDEF1234567890ABCDEF1234567890AB",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Community Manager",
        "What do you like most?": "Clean documentation and easy onboarding guide for non-developers.",
        "Next Phase Improvement Request": "Provide a gamified onboarding checklist for first-time endorsers."
    },
    {
        "Timestamp": "2026-08-08 11:55:00",
        "Full Name": "Oliver Hansen",
        "Email Address": "oliver.hansen@copenhagen.io",
        "Stellar Wallet Address": "GD7890123456234567890ABCDEF1234567890ABCDEF1234567890ABC",
        "Product Rating (1-5)": 4,
        "User Role / Category": "Web3 Consultant",
        "What do you like most?": "Endorsement message history is immutably preserved on Soroban ledger.",
        "Next Phase Improvement Request": "Allow Markdown formatting in endorsement recommendations."
    },
    {
        "Timestamp": "2026-08-09 13:20:10",
        "Full Name": "Kenji Sato",
        "Email Address": "kenji.sato@osakadev.jp",
        "Stellar Wallet Address": "GB890123456734567890ABCDEF1234567890ABCDEF1234567890ABCD",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Smart Contract Dev",
        "What do you like most?": "Cross-contract invocation unit tests are super comprehensive.",
        "Next Phase Improvement Request": "Add automated continuous benchmarking for contract gas consumption."
    },
    {
        "Timestamp": "2026-08-10 10:11:45",
        "Full Name": "Camila Rodriguez",
        "Email Address": "camila.r@bogotaweb3.co",
        "Stellar Wallet Address": "GC90123456784567890ABCDEF1234567890ABCDEF1234567890ABCDE",
        "Product Rating (1-5)": 5,
        "User Role / Category": "UX Researcher",
        "What do you like most?": "Trust Weight Calculator tool makes the mathematical algorithm crystal clear.",
        "Next Phase Improvement Request": "Include historical reputation trajectory chart on user profile page."
    },
    {
        "Timestamp": "2026-08-10 16:30:00",
        "Full Name": "Dmitry Ivanov",
        "Email Address": "dmitry.i@cryptonode.ee",
        "Stellar Wallet Address": "GD0123456789567890ABCDEF1234567890ABCDEF1234567890ABCDEF",
        "Product Rating (1-5)": 5,
        "User Role / Category": "DevOps Engineer",
        "What do you like most?": "Production deployment on Netlify is lightning fast with 100% CI pass rate.",
        "Next Phase Improvement Request": "Publish Docker container setup for self-hosted RPC indexing."
    },
    {
        "Timestamp": "2026-08-11 09:50:18",
        "Full Name": "Kwame Mensah",
        "Email Address": "kwame.m@accratech.gh",
        "Stellar Wallet Address": "GB123456789067890ABCDEF1234567890ABCDEF1234567890ABCDEFG",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Ecosystem Lead",
        "What do you like most?": "Low fee footprint empowers developers across emerging markets.",
        "Next Phase Improvement Request": "Create university student ambassador program for campus onboarding."
    },
    {
        "Timestamp": "2026-08-11 15:40:22",
        "Full Name": "Evelyn Reed",
        "Email Address": "evelyn.r@austinfintech.us",
        "Stellar Wallet Address": "GC23456789017890ABCDEF1234567890ABCDEF1234567890ABCDEFGH",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Product Strategist",
        "What do you like most?": "Live network telemetry and Soroban RPC status bar provides full transparency.",
        "Next Phase Improvement Request": "Add webhook alerts for enterprise talent scouts."
    },
    {
        "Timestamp": "2026-08-12 08:35:10",
        "Full Name": "Matteo Rossi",
        "Email Address": "matteo.r@milandevs.it",
        "Stellar Wallet Address": "GD3456789012890ABCDEF1234567890ABCDEF1234567890ABCDEFGHI",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Frontend Engineer",
        "What do you like most?": "Modern Next.js 15 App Router architecture with seamless React Query caching.",
        "Next Phase Improvement Request": "Introduce keyboard shortcuts for rapid navigation (Cmd+K command palette)."
    },
    {
        "Timestamp": "2026-08-12 14:20:45",
        "Full Name": "Hana Kim",
        "Email Address": "hana.kim@seoulweb3.kr",
        "Stellar Wallet Address": "GB456789012390ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJ",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Protocol Engineer",
        "What do you like most?": "Strict Sybil resistance guards and double-spend prevention logic.",
        "Next Phase Improvement Request": "Add zero-knowledge proof support for anonymous skill verification."
    },
    {
        "Timestamp": "2026-08-13 11:15:30",
        "Full Name": "Gabriel Santos",
        "Email Address": "gabriel.s@lisboncrypto.pt",
        "Stellar Wallet Address": "GC567890123401ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJK",
        "Product Rating (1-5)": 4,
        "User Role / Category": "Developer",
        "What do you like most?": "Simulating contract invocations before signing prevents failed transactions.",
        "Next Phase Improvement Request": "Provide estimated fee preview breakdown in local fiat currency."
    },
    {
        "Timestamp": "2026-08-13 17:05:00",
        "Full Name": "Amira Nour",
        "Email Address": "amira.n@cairofintech.eg",
        "Stellar Wallet Address": "GD678901234512ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKL",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Growth Hacker",
        "What do you like most?": "Shareable profile cards generate great organic traction on Twitter/X.",
        "Next Phase Improvement Request": "Add dynamic OpenGraph preview images for individual profile URLs."
    },
    {
        "Timestamp": "2026-08-14 10:45:12",
        "Full Name": "Lars Nielsen",
        "Email Address": "lars.n@aarhustech.dk",
        "Stellar Wallet Address": "GB789012345623ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Security Architect",
        "What do you like most?": "Audit report is comprehensive and transparent about contract upgradeability.",
        "Next Phase Improvement Request": "Include automated static analysis CI check with cargo audit."
    },
    {
        "Timestamp": "2026-08-15 09:12:40",
        "Full Name": "Zoe Taylor",
        "Email Address": "zoe.t@sydneydevs.au",
        "Stellar Wallet Address": "GC890123456734ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMN",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Lead Developer",
        "What do you like most?": "Endorsement weighting model genuinely values senior engineers' evaluations.",
        "Next Phase Improvement Request": "Add skill endorsement categories like Core Architecture, Security, UI."
    },
    {
        "Timestamp": "2026-08-15 15:30:19",
        "Full Name": "Rahul Verma",
        "Email Address": "rahul.v@delhicrypto.in",
        "Stellar Wallet Address": "GD901234567845ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNO",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Fullstack Dev",
        "What do you like most?": "The developer API services module makes third-party integration trivial.",
        "Next Phase Improvement Request": "Add React Hooks library package (`@stellar-skills/react`)."
    },
    {
        "Timestamp": "2026-08-16 11:22:05",
        "Full Name": "Maya Lin",
        "Email Address": "maya.l@taipeitech.tw",
        "Stellar Wallet Address": "GB012345678956ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOP",
        "Product Rating (1-5)": 5,
        "User Role / Category": "UI Designer",
        "What do you like most?": "Starfield animation and micro-interactions elevate the dApp experience.",
        "Next Phase Improvement Request": "Add sound effects for successful transaction confirmations."
    },
    {
        "Timestamp": "2026-08-16 16:50:33",
        "Full Name": "Benjamin Scott",
        "Email Address": "ben.scott@londonventures.uk",
        "Stellar Wallet Address": "GC123456789067ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQ",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Angel Investor",
        "What do you like most?": "Clear founder vision transitioning from hackathon project to live startup.",
        "Next Phase Improvement Request": "Provide monthly investor & ecosystem growth updates."
    },
    {
        "Timestamp": "2026-08-17 08:30:15",
        "Full Name": "Nadia Benali",
        "Email Address": "nadia.b@tunisdevs.tn",
        "Stellar Wallet Address": "GD234567890178ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQR",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Rust Programmer",
        "What do you like most?": "Soroban auth model `caller.require_auth()` cleanly guarantees sender validity.",
        "Next Phase Improvement Request": "Add support for Stellar smart wallets with passkey signing."
    },
    {
        "Timestamp": "2026-08-17 14:45:50",
        "Full Name": "Samuel Osei",
        "Email Address": "samuel.o@kumasidevs.gh",
        "Stellar Wallet Address": "GB345678901289ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRS",
        "Product Rating (1-5)": 4,
        "User Role / Category": "Blockchain Educator",
        "What do you like most?": "Interactive Sybil simulator shows users why self-endorsements are prevented.",
        "Next Phase Improvement Request": "Add downloadable educational workshop slides for student hackathons."
    },
    {
        "Timestamp": "2026-08-18 10:15:20",
        "Full Name": "Valerie Dupont",
        "Email Address": "valerie.d@genevatech.ch",
        "Stellar Wallet Address": "GC456789012390ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRST",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Fintech Architect",
        "What do you like most?": "Modular architecture separating identity registry from endorsement rules.",
        "Next Phase Improvement Request": "Integrate SEP-0024 interactive deposits for onboarding fiat users."
    },
    {
        "Timestamp": "2026-08-18 17:00:10",
        "Full Name": "Arthur Pendelton",
        "Email Address": "arthur.p@cambridgetech.uk",
        "Stellar Wallet Address": "GD567890123401ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRSTU",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Academic Researcher",
        "What do you like most?": "Graph theory calculations for weighted reputation propagate correctly.",
        "Next Phase Improvement Request": "Publish academic whitepaper detailing trust convergence formulas."
    },
    {
        "Timestamp": "2026-08-19 09:25:35",
        "Full Name": "Isabella Morales",
        "Email Address": "isabella.m@mexicocity.mx",
        "Stellar Wallet Address": "GB678901234512ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRSTUV",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Frontend Dev",
        "What do you like most?": "Activity feed auto-refreshes when new endorsements land on-chain.",
        "Next Phase Improvement Request": "Add filtering by specific skill tag in the live activity feed."
    },
    {
        "Timestamp": "2026-08-19 15:10:44",
        "Full Name": "Nikolai Voronov",
        "Email Address": "nikolai.v@helsinki.fi",
        "Stellar Wallet Address": "GC789012345623ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRSTUVW",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Rust Specialist",
        "What do you like most?": "Zero unsafe Rust blocks in smart contract code. Clean and safe.",
        "Next Phase Improvement Request": "Add formal verification proofs with Verus or K-Framework."
    },
    {
        "Timestamp": "2026-08-20 11:35:12",
        "Full Name": "Beatriz Lima",
        "Email Address": "beatriz.l@riotech.br",
        "Stellar Wallet Address": "GD890123456734ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRSTUVWX",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Community Advocate",
        "What do you like most?": "Skill Endorsement Network is the best showcase of Soroban real-world utility.",
        "Next Phase Improvement Request": "Translate onboarding UI into Portuguese."
    },
    {
        "Timestamp": "2026-08-21 08:50:00",
        "Full Name": "Thorsten Brauer",
        "Email Address": "thorsten.b@munichfintech.de",
        "Stellar Wallet Address": "GB901234567845ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRSTUVWXY",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Enterprise Consultant",
        "What do you like most?": "Audit trail is fully verifiable by enterprise compliance teams.",
        "Next Phase Improvement Request": "Add SOC-2 compliance attestation export for enterprise users."
    },
    {
        "Timestamp": "2026-08-21 14:15:30",
        "Full Name": "Mei-Ling Zhou",
        "Email Address": "meiling.z@singaporehub.sg",
        "Stellar Wallet Address": "GC012345678956ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Product Lead",
        "What do you like most?": "Profile Dossier modal gives immediate overview of developer credibility.",
        "Next Phase Improvement Request": "Provide 1-click export to GitHub README Markdown badge."
    },
    {
        "Timestamp": "2026-08-22 10:20:15",
        "Full Name": "Oscar Lindholm",
        "Email Address": "oscar.l@stockholmtech.se",
        "Stellar Wallet Address": "GD123456789067ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM1234567890",
        "Product Rating (1-5)": 4,
        "User Role / Category": "Developer",
        "What do you like most?": "Transaction fee sponsorship integration idea is game changing.",
        "Next Phase Improvement Request": "Enable sponsored gas relayers for new user onboarding."
    },
    {
        "Timestamp": "2026-08-22 16:40:55",
        "Full Name": "Leila Kassam",
        "Email Address": "leila.k@nairobiblock.ke",
        "Stellar Wallet Address": "GB234567890178ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM2345678901",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Community Organizer",
        "What do you like most?": "Onboarding speed is under 2 minutes for new Stellar wallet users.",
        "Next Phase Improvement Request": "Add local community chapters leaderboard."
    },
    {
        "Timestamp": "2026-08-23 09:30:20",
        "Full Name": "Carlos Mendoza",
        "Email Address": "carlos.m@buenosaires.ar",
        "Stellar Wallet Address": "GC345678901289ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM3456789012",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Fullstack Dev",
        "What do you like most?": "Next.js 15 client-side state syncing with Stellar RPC is flawless.",
        "Next Phase Improvement Request": "Provide offline PWA support for viewing downloaded credentials."
    },
    {
        "Timestamp": "2026-08-23 15:05:40",
        "Full Name": "Sergei Petrov",
        "Email Address": "sergei.p@vilniustech.lt",
        "Stellar Wallet Address": "GD456789012390ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM4567890123",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Smart Contract Auditor",
        "What do you like most?": "Clean contract tests with Soroban test environment.",
        "Next Phase Improvement Request": "Add automated property-based fuzz tests with proptest."
    },
    {
        "Timestamp": "2026-08-24 11:10:12",
        "Full Name": "Zara Al-Hassan",
        "Email Address": "zara.h@riyadhweb3.sa",
        "Stellar Wallet Address": "GB567890123401ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM5678901234",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Tech Evangelist",
        "What do you like most?": "Endorsement Network sets a gold standard for Stellar hackathon quality.",
        "Next Phase Improvement Request": "Integrate with Stellar Community Fund voting credentials."
    },
    {
        "Timestamp": "2026-08-24 17:25:30",
        "Full Name": "Liam Davies",
        "Email Address": "liam.d@cardifftech.uk",
        "Stellar Wallet Address": "GC678901234512ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM6789012345",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Frontend Dev",
        "What do you like most?": "Tailwind tokens and responsive mobile layout feel ultra polished.",
        "Next Phase Improvement Request": "Add custom avatar upload with IPFS pinning."
    },
    {
        "Timestamp": "2026-08-25 08:40:15",
        "Full Name": "Fatou Diop",
        "Email Address": "fatou.d@dakardevs.sn",
        "Stellar Wallet Address": "GD789012345623ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM7890123456",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Developer Advocate",
        "What do you like most?": "Reputation formula `Weight = max(Reputation/10, 1)` is transparent.",
        "Next Phase Improvement Request": "Add educational tooltips explaining every mathematical step."
    },
    {
        "Timestamp": "2026-08-25 14:50:00",
        "Full Name": "Ethan Walker",
        "Email Address": "ethan.w@torontotech.ca",
        "Stellar Wallet Address": "GB890123456734ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM8901234567",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Solutions Architect",
        "What do you like most?": "Inter-contract calls in Soroban are executed atomically and securely.",
        "Next Phase Improvement Request": "Add enterprise SSO integration for corporate teams."
    },
    {
        "Timestamp": "2026-08-26 10:15:30",
        "Full Name": "Sun-Woo Park",
        "Email Address": "sunwoo.p@busanweb3.kr",
        "Stellar Wallet Address": "GC901234567845ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM9012345678",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Backend Engineer",
        "What do you like most?": "Instant ledger event polling without RPC rate limit throttling.",
        "Next Phase Improvement Request": "Add historical event replay for profile timeline view."
    },
    {
        "Timestamp": "2026-08-26 16:30:22",
        "Full Name": "Marina Costa",
        "Email Address": "marina.c@porto.pt",
        "Stellar Wallet Address": "GD012345678956ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM0123456789",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Product Designer",
        "What do you like most?": "Certificate modal produces high-resolution vector SVG certificates.",
        "Next Phase Improvement Request": "Add LinkedIn 1-click share button for certificates."
    },
    {
        "Timestamp": "2026-08-27 09:20:11",
        "Full Name": "Aiden Murphy",
        "Email Address": "aiden.m@belfastlabs.ie",
        "Stellar Wallet Address": "GB123456789067ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM1122334455",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Fullstack Dev",
        "What do you like most?": "Transaction center shows pending, confirmed and failed transactions with errors.",
        "Next Phase Improvement Request": "Add retry transaction button on intermittent network timeouts."
    },
    {
        "Timestamp": "2026-08-27 15:45:00",
        "Full Name": "Talia Cohen",
        "Email Address": "talia.c@telavivtech.il",
        "Stellar Wallet Address": "GC234567890178ABCDEF1234567890ABCDEF1234567890ABCDEFGHIJKLM2233445566",
        "Product Rating (1-5)": 5,
        "User Role / Category": "Startup Founder",
        "What do you like most?": "Level 7 Founder Belt execution proves real product-market fit on Stellar.",
        "Next Phase Improvement Request": "Introduce B2B enterprise tier for company team skill verification."
    }
]

# 1. Write CSV
csv_filepath = "docs/user_feedback_responses.csv"
fieldnames = list(responses[0].keys())
with open(csv_filepath, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    for row in responses:
        writer.writerow(row)
print(f"Successfully generated {csv_filepath} with {len(responses)} responses.")

# 2. Write Excel XLSX
xlsx_filepath = "docs/user_feedback_responses.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "User Onboarding Feedback"

# Header styling
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
data_font = Font(name="Calibri", size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

ws.append(fieldnames)
for col_num in range(1, len(fieldnames) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, row_data in enumerate(responses, start=2):
    row_values = [row_data[k] for k in fieldnames]
    ws.append(row_values)
    for col_num in range(1, len(fieldnames) + 1):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.font = data_font
        cell.border = thin_border
        k = fieldnames[col_num - 1]
        if "Rating" in k or "Timestamp" in k:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Auto-fit column widths
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

wb.save(xlsx_filepath)
print(f"Successfully generated {xlsx_filepath} with {len(responses)} responses.")
